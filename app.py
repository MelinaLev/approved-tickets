import os
from functools import wraps
from flask import Flask, render_template, request, send_file, flash, redirect, url_for, Response
import pandas as pd
from io import BytesIO
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import tempfile

app = Flask(__name__)
app.secret_key = os.environ.get("FLASK_SECRET_KEY", "dev-secret-change-me")

# --- Basic Auth (set APP_USER and APP_PASS in Render env vars) ---
USERNAME = os.getenv("APP_USER")
PASSWORD = os.getenv("APP_PASS")

def requires_auth(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if not USERNAME or not PASSWORD:  # no creds set -> skip auth
            return f(*args, **kwargs)
        auth = request.authorization
        if not auth or auth.username != USERNAME or auth.password != PASSWORD:
            return Response("Login required", 401, {"WWW-Authenticate": 'Basic realm="Login"'})
        return f(*args, **kwargs)
    return decorated

ALLOWED_EXTENSIONS = {"csv", "xlsx", "xls"}
app.config["MAX_CONTENT_LENGTH"] = 50 * 1024 * 1024  # 50 MB

def allowed_file(filename):
    return "." in filename and filename.rsplit(".", 1)[1].lower() in ALLOWED_EXTENSIONS

def read_table(file_storage):
    name = file_storage.filename.lower()
    if name.endswith(".csv"):
        return pd.read_csv(file_storage)
    else:
        return pd.read_excel(file_storage)

def extract_tickets(cell):
    if pd.isna(cell):
        return []
    raw = str(cell).replace(";", ",")
    parts = [p.strip() for p in raw.split(",")]
    tickets = []
    for p in parts:
        tickets.extend([pp for pp in p.split() if pp])
    return [t.strip() for t in tickets if t.strip()]

def compute_status(main_df, approved_df, tickets_col="Tickets", approved_ticket_col="Ticket"):
    cols_lower = {c.lower(): c for c in approved_df.columns}
    status_col = cols_lower.get("ticket status")
    approved_series = approved_df[approved_ticket_col].astype(str).str.strip()
    if status_col:
        mask = approved_df[status_col].astype(str).str.strip().str.lower() == "approved"
        approved_set = set(approved_series[mask])
    else:
        approved_set = set(approved_series)

    statuses, missing_list = [], []
    for _, row in main_df.iterrows():
        tickets = extract_tickets(row.get(tickets_col))
        if not tickets:
            statuses.append("No Tickets")
            missing_list.append("")
            continue
        hits = [t for t in tickets if t in approved_set]
        missing = [t for t in tickets if t not in approved_set]
        if len(hits) == len(tickets):
            statuses.append("Ready for Payment")
        elif hits:
            statuses.append("Pending")
        else:
            statuses.append("Not Approved")
        missing_list.append(", ".join(missing))

    out = main_df.copy()
    out["Approval Status"] = statuses
    out["Missing Tickets"] = missing_list
    return out

def add_highlighting_and_summary(xlsx_path):
    wb = load_workbook(xlsx_path)
    ws = wb.active
    headers = {ws.cell(1, i).value: i for i in range(1, ws.max_column + 1)}
    status_idx = headers.get("Approval Status")

    green = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    yellow = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    red = PatternFill(start_color="F2DCDB", end_color="F2DCDB", fill_type="solid")
    gray = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")

    counts = {"Ready for Payment":0, "Pending":0, "Not Approved":0, "No Tickets":0}
    if status_idx:
        for r in range(2, ws.max_row + 1):
            val = ws.cell(r, status_idx).value
            if val == "Ready for Payment":
                ws.cell(r, status_idx).fill = green
            elif val == "Pending":
                ws.cell(r, status_idx).fill = yellow
            elif val == "Not Approved":
                ws.cell(r, status_idx).fill = red
            elif val == "No Tickets":
                ws.cell(r, status_idx).fill = gray
            if val in counts:
                counts[val] += 1

    s = wb.create_sheet("Summary")
    s.append(["Status", "Count"])
    for k, v in counts.items():
        s.append([k, v])
    wb.save(xlsx_path)

@app.route("/", methods=["GET", "POST"])
@requires_auth
def index():
    if request.method == "POST":
        main = request.files.get("main_report")
        approved = request.files.get("approved")
        tickets_col = (request.form.get("tickets_col") or "Tickets").strip() or "Tickets"
        approved_ticket_col = (request.form.get("approved_ticket_col") or "Ticket").strip() or "Ticket"

        if not main or not approved:
            flash("Please upload both files.")
            return redirect(url_for("index"))
        if not (allowed_file(main.filename) and allowed_file(approved.filename)):
            flash("Files must be CSV/XLSX.")
            return redirect(url_for("index"))

        try:
            main_df = read_table(main)
            approved_df = read_table(approved)
            main_df.columns = [c.strip() for c in main_df.columns]
            approved_df.columns = [c.strip() for c in approved_df.columns]
            if tickets_col not in main_df.columns:
                raise ValueError(f"Column '{tickets_col}' not found in Main report.")
            if approved_ticket_col not in approved_df.columns:
                raise ValueError(f"Column '{approved_ticket_col}' not found in Approved file.")

            out_df = compute_status(main_df, approved_df, tickets_col, approved_ticket_col)
            with tempfile.TemporaryDirectory() as td:
                xlsx_path = os.path.join(td, "Main_report_with_approvals.xlsx")
                out_df.to_excel(xlsx_path, index=False)
                add_highlighting_and_summary(xlsx_path)
                with open(xlsx_path, "rb") as f:
                    data = f.read()
            return send_file(BytesIO(data),
                             download_name="Main_report_with_approvals.xlsx",
                             as_attachment=True,
                             mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        except Exception as e:
            flash(f"Error: {e}")
            return redirect(url_for("index"))
    return render_template("index.html")

if __name__ == "__main__":
    port = int(os.environ.get("PORT", 5000))
    app.run(host="0.0.0.0", port=port, debug=True)